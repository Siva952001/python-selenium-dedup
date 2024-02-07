from fastapi import FastAPI, Form, Request
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from fuzzywuzzy import fuzz

app = FastAPI()
templates = Jinja2Templates(directory="templates")

# Load data from Excel file
def load_excel_data(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append({
            "first_name": row[0],
            "last_name": row[1],
            "dob": row[2],
            "gender":row[3],
            "email": row[4],
            "address": row[5]
        })
    return data


# Endpoint to display form
@app.get("/test", response_class=HTMLResponse)
async def show_form(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})



# Endpoint to handle form submission
@app.post("/check_duplicates/", response_class=HTMLResponse)
async def check_duplicates(request: Request,
                           first_name: str = Form(...),
                           last_name: str = Form(...),
                           gender: str = Form(...),
                           dob: str = Form(...),
                           email: str = Form(...),
                           address: str = Form(...)):
    form_data = {
        "first_name": first_name,
        "last_name": last_name,
        "gender": gender,
        "dob": dob,
        "email": email,
        "address": address
    }

    # Find duplicates
    duplicates = []
    total_similarity = 0
    num_similarities = 0

    excel_data = load_excel_data("modified_dedup.xlsx")

    for row in excel_data:
        first_name_similarity = fuzz.token_sort_ratio(str(row["first_name"]), str(first_name))
        last_name_similarity = fuzz.token_sort_ratio(str(row["last_name"]), str(last_name))
        gender_similarity = fuzz.token_sort_ratio(str(row["dob"]), str(dob))
        dob_similarity = fuzz.token_sort_ratio(str(row["gender"]), str(gender))
        email_similarity = fuzz.token_sort_ratio(str(row["email"]), str(email))
        address_similarity=fuzz.token_sort_ratio(str(row["address"]), str(address))
        
        similarity_values=[]
        similarity_values.extend([first_name_similarity, last_name_similarity, dob_similarity,gender_similarity, email_similarity, address_similarity])

        print(similarity_values,sum(similarity_values)/len(similarity_values))

        if (sum(similarity_values)/len(similarity_values)) > 75 :
            duplicates.append({
                    "record": row,
                    "similarity_percentage": similarity_values
                })

    return templates.TemplateResponse("results.html", {
        "request": request,
        "form_data": form_data,
        "duplicates": duplicates,
        "percentage": similarity_values
    })

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="192.168.2.251", port=8000)
