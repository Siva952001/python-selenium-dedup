from openpyxl import load_workbook, Workbook
import datetime

# Load data from Excel file
workbook = load_workbook("dedup.xlsx")
sheet = workbook.active
data = []

# Iterate through rows and normalize date if needed
for row in sheet.iter_rows(min_row=2, values_only=True):
    if isinstance(row[2], datetime.datetime):
        # Format datetime object as string in 'DD-MM-YYYY' format
        row_list = list(row)
        row_list[2] = row[2].strftime('%d-%m-%Y')
        data.append(tuple(row_list))
    else:
        # If the date is already a string, append it directly
        data.append(row)

# Create a new workbook and select the active sheet
new_workbook = Workbook()
new_sheet = new_workbook.active

# Add the modified data to the new sheet
for row_index, row_data in enumerate(data, start=1):
    for column_index, cell_value in enumerate(row_data, start=1):
        new_sheet.cell(row=row_index, column=column_index, value=cell_value)

# Save the workbook to a new Excel file
new_workbook.save("modified_dedup.xlsx")
