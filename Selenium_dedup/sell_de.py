import time 
import pytesseract
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



import subprocess 
options = Options()
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
import xlsxwriter

excel_file_path = r'D:\Selenium_dedup\dedup.xlsx'
sheet_name = 'Test'

# Load the Excel workbook and get the active sheet
workbook = load_workbook(excel_file_path)
sheet = workbook[sheet_name]

# Create an new Excel file and add a worksheet.
workbook = xlsxwriter.Workbook('dedup_testing.xlsx')

worksheet= workbook.add_worksheet()

driver.get("http://192.168.2.251:8000/test")    

row_headings = ['fname_actual', 'fnamegiven', 'Accuracy']
bold_format = workbook.add_format({'bold': True, 'font_size': 10})

worksheet.write_row(0, 0, row_headings,bold_format)

col=['A','B','C']

# for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
#     print("first_name:", row[0], "Last_name:", row[1])

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
    time.sleep(3)
    driver.find_element(By.ID, "first_name").send_keys(row[0])
    driver.find_element(By.ID, "last_name").send_keys(row[1])
    driver.find_element(By.ID, "dob").send_keys(row[2])
    driver.find_element(By.ID, "gender").send_keys(row[3])
    driver.find_element(By.ID, "email").send_keys(row[4])
    driver.find_element(By.ID, "address").send_keys(row[5])
    driver.find_element(By.LINK_TEXT, "Check Duplicates").click()
    


workbook.close()
driver.close()