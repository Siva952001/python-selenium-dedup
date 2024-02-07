from openpyxl import load_workbook

# Assuming your Excel file has a sheet named 'Test'
excel_file_path = r'D:\Selenium_dedup\dedup.xlsx'
sheet_name = 'Test'

# Load the Excel workbook and get the active sheet
workbook = load_workbook(excel_file_path)
sheet = workbook[sheet_name]

# Iterate through rows and print values
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
    print("first_name:", row[0], "Last_name:", row[1])

# Close the Excel workbook
workbook.close()