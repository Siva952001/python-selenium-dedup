import time 
import random 
import string
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook, load_workbook
import datetime

options = Options()
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

excel_file_path = r'dedup.xlsx'
sheet_name = 'Test'

# Load the Excel workbook and get the active sheet by name
workbook = load_workbook(excel_file_path)
try:
    sheet = workbook[sheet_name]
except KeyError:
    print(f"Sheet '{sheet_name}' not found in the workbook.")
    workbook.close()
    driver.quit()
    exit()

# Create a new workbook and worksheet
new_workbook = Workbook()
new_worksheet = new_workbook.active

row_headings = ['S.no', 'Modified_f_name','Actual_f_name' ,'Modified_l_name','Actual_l_name', 'Modified_email','Actual_email', 'Modified_address','Actual_address','Modified_gender','Actual_gender','Modified_dob','Actual_dob']
new_worksheet.append(row_headings)

driver.get("http://192.168.2.251:8000/test")

for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=1):
    time.sleep(3)

    # Generate random modification for first name
    random_alphabet_first = ''.join(random.choices(string.ascii_letters, k=random.choice([1, 2])))
    first_name_length = len(row[0])
    middle_index_first = first_name_length // 2
    modified_first_name = row[0][:middle_index_first] + random_alphabet_first.lower() + row[0][middle_index_first:]

    # Generate random modification for last name
    random_alphabet_last = ''.join(random.choices(string.ascii_letters, k=random.choice([1, 2])))
    last_name_length = len(row[1])
    middle_index_last = last_name_length // 2
    modified_last_name = row[1][:middle_index_last] + random_alphabet_last.lower() + row[1][middle_index_last:]

    # Generate random modification for email
    middle_index_email = len(row[4]) // 2
    modified_email = row[4][:middle_index_email] + random.choice(string.ascii_lowercase) + row[4][middle_index_email:]

    # Generate random modification for address
    middle_index_address = len(row[5]) // 2
    modified_address = row[5][:middle_index_address] + ''.join(random.choices(string.ascii_lowercase, k=3)) + row[5][middle_index_address:]
    
     # Randomly select "Male" or "Female"
    gender = random.choice(["Male", "Female"])
    
    # Check if row[2] is a datetime.datetime object
    if isinstance(row[2], datetime.datetime):
        # Convert the DOB to a string
        dob_string = row[2].strftime('%d-%m-%Y')
    else:
        # If row[2] is already a string, use it directly
        dob_string = row[2]

    # Extract the day, month, and year components from the DOB string
    day, month, year = map(int, dob_string.split('-'))

    # Generate random two-digit number
    random_number = str(random.randint(10, 99))

    # Modify the last two digits of the year component
    modified_year = str(year)[:-2] + random_number

    # Reconstruct the modified DOB string
    modified_dob = f"{day:02d}-{month:02d}-{modified_year}"


    # Fill in the form fields
    driver.find_element(By.ID, "first_name").clear()
    driver.find_element(By.ID, "first_name").send_keys(modified_first_name)
    driver.find_element(By.ID, "last_name").send_keys(modified_last_name)
    if isinstance(row[2], str):
        driver.find_element(By.ID, "dob").send_keys(row[2])
    elif isinstance(row[2], datetime.datetime):
        dob_string = row[2].strftime('%d-%m-%Y')
        driver.find_element(By.ID, "dob").send_keys(modified_dob)
    driver.find_element(By.ID, "gender").send_keys(gender)
    driver.find_element(By.ID, "email").send_keys(modified_email)
    driver.find_element(By.ID, "address").send_keys(modified_address)
    driver.find_element(By.XPATH, "/html/body/main/form/button").click()
    time.sleep(3)
    
    # Assuming the result is obtained from the webpage (replace this with your actual logic)
    result_from_f_name = driver.find_element(By.XPATH, "/html/body/table[2]/tbody/tr/td[1]").text
    result_from_l_name = driver.find_element(By.XPATH, "/html/body/table[2]/tbody/tr/td[2]").text
    result_from_gender = driver.find_element(By.XPATH, "/html/body/table[2]/tbody/tr/td[3]").text
    result_from_dob = driver.find_element(By.XPATH, "/html/body/table[2]/tbody/tr/td[4]").text
    result_from_email = driver.find_element(By.XPATH, "/html/body/table[2]/tbody/tr/td[5]").text
    result_from_address = driver.find_element(By.XPATH, "/html/body/table[2]/tbody/tr/td[6]").text
    
    # Write the result to the new worksheet
    new_worksheet.append([row_number,modified_first_name,result_from_f_name,modified_last_name,result_from_l_name,modified_email,result_from_email,modified_address,result_from_address,gender,result_from_gender,modified_dob,result_from_dob])
    #Save the new workbook
    new_workbook.save('dedup_testing_with_results.xlsx')
    print("New workbook saved successfully.")
    
    driver.find_element(By.XPATH, "/html/body/header/nav/div/a").click()
    time.sleep(1)

# Close the workbooks and the WebDriver
workbook.close()
new_workbook.close()
driver.quit()
